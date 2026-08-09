"""Deterministic CSV and Excel exports for the future DEMATEL engine."""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO
from typing import Literal

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import CANNOT_ASSESS_VALUE, FACTOR_CODES, load_factor_catalogue
from hierarchical_questionnaire import (
    all_hierarchical_relationships,
    get_matrix_definition,
    matrix_definitions,
)
from models import (
    AssignmentRecord,
    DistributedResponseRecord,
    HierarchicalQuestionnaireRecord,
    HierarchicalResponseRecord,
    ResponseRecord,
)
from questionnaire_sets import all_relationships

LONG_COLUMNS = [
    "submission_id",
    "expert_code",
    "timestamp",
    "from_factor",
    "to_factor",
    "linguistic_value",
    "tfn_l",
    "tfn_m",
    "tfn_u",
    "is_diagonal",
]


@dataclass(frozen=True, slots=True)
class ExportBundle:
    """In-memory files generated for one submitted expert matrix."""

    long_csv: bytes
    wide_csv: bytes
    long_excel: bytes
    wide_excel: bytes


@dataclass(frozen=True, slots=True)
class AdministratorExportBundle:
    """Combined distributed-response files for the research administrator."""

    responses_csv: bytes
    complete_excel: bytes


def records_to_long_dataframe(
    records: Sequence[ResponseRecord],
) -> pd.DataFrame:
    """Return a stable row-major long-format dataframe."""

    frame = pd.DataFrame.from_records(records, columns=LONG_COLUMNS)
    expected_rows = len(FACTOR_CODES) ** 2
    if len(frame) != expected_rows:
        raise ValueError(f"Expected {expected_rows} records; received {len(frame)}.")

    factor_order = {factor: index for index, factor in enumerate(FACTOR_CODES)}
    frame["_from_order"] = frame["from_factor"].map(factor_order)
    frame["_to_order"] = frame["to_factor"].map(factor_order)
    if frame[["_from_order", "_to_order"]].isna().any().any():
        raise ValueError("Records contain an unknown factor code.")
    frame = frame.sort_values(["_from_order", "_to_order"]).drop(
        columns=["_from_order", "_to_order"]
    )
    return frame.reset_index(drop=True)


def records_to_wide_dataframe(
    records: Sequence[ResponseRecord],
    value: Literal["linguistic_value", "tfn_l", "tfn_m", "tfn_u"],
) -> pd.DataFrame:
    """Pivot one canonical value into an ordered 18×18 wide matrix."""

    long_frame = records_to_long_dataframe(records)
    if long_frame.duplicated(["from_factor", "to_factor"]).any():
        raise ValueError("Duplicate ordered factor pairs cannot be exported.")

    wide = long_frame.pivot(
        index="from_factor", columns="to_factor", values=value
    ).reindex(index=FACTOR_CODES, columns=FACTOR_CODES)
    if wide.isna().any().any():
        raise ValueError("The wide matrix is missing one or more factor pairs.")
    wide.index.name = "from_factor"
    wide.columns.name = "to_factor"
    return wide


def _definitions_dataframe() -> pd.DataFrame:
    return pd.DataFrame.from_records(
        {
            "factor_code": item.code,
            "dimension": item.dimension,
            "criterion": item.criterion,
            "full_definition": item.definition,
        }
        for item in load_factor_catalogue()
    )


def _metadata_dataframe(long_frame: pd.DataFrame) -> pd.DataFrame:
    first = long_frame.iloc[0]
    return pd.DataFrame(
        {
            "field": [
                "submission_id",
                "expert_code",
                "timestamp",
                "matrix_orientation",
                "off_diagonal_comparisons",
                "diagonal_rule",
            ],
            "value": [
                first["submission_id"],
                first["expert_code"],
                first["timestamp"],
                "ROW factor influences COLUMN factor",
                "306",
                "Fixed zero TFN (0.00, 0.00, 0.00)",
            ],
        }
    )


def _format_workbook(writer: pd.ExcelWriter) -> None:
    """Apply restrained, machine-safe formatting to every workbook sheet."""

    header_fill = PatternFill("solid", fgColor="163A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in worksheet.columns:
            max_length = min(
                60,
                max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                + 2,
            )
            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max(10, max_length)


def _dataframe_to_csv_bytes(frame: pd.DataFrame, *, index: bool) -> bytes:
    return frame.to_csv(index=index, lineterminator="\n").encode("utf-8-sig")


def _long_excel_bytes(long_frame: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        long_frame.to_excel(writer, sheet_name="Long_Data", index=False)
        _metadata_dataframe(long_frame).to_excel(
            writer, sheet_name="Metadata", index=False
        )
        _definitions_dataframe().to_excel(
            writer, sheet_name="Factor_Definitions", index=False
        )
        _format_workbook(writer)
    return buffer.getvalue()


def _wide_excel_bytes(
    long_frame: pd.DataFrame, records: Sequence[ResponseRecord]
) -> bytes:
    buffer = BytesIO()
    sheet_map = {
        "Linguistic": "linguistic_value",
        "TFN_L": "tfn_l",
        "TFN_M": "tfn_m",
        "TFN_U": "tfn_u",
    }
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, value_column in sheet_map.items():
            records_to_wide_dataframe(records, value_column).to_excel(
                writer, sheet_name=sheet_name, index=True
            )
        _metadata_dataframe(long_frame).to_excel(
            writer, sheet_name="Metadata", index=False
        )
        _definitions_dataframe().to_excel(
            writer, sheet_name="Factor_Definitions", index=False
        )
        _format_workbook(writer)
    return buffer.getvalue()


def generate_exports(records: Sequence[ResponseRecord]) -> ExportBundle:
    """Generate CSV and Excel files in both long and wide layouts."""

    long_frame = records_to_long_dataframe(records)
    linguistic_wide = records_to_wide_dataframe(records, "linguistic_value")
    return ExportBundle(
        long_csv=_dataframe_to_csv_bytes(long_frame, index=False),
        wide_csv=_dataframe_to_csv_bytes(linguistic_wide, index=True),
        long_excel=_long_excel_bytes(long_frame),
        wide_excel=_wide_excel_bytes(long_frame, records),
    )


ADMIN_RESPONSE_COLUMNS = [
    "respondent_id",
    "expert_code",
    "set_id",
    "source_variable_code",
    "source_variable_name",
    "target_variable_code",
    "target_variable_name",
    "linguistic_response",
    "tfn_l",
    "tfn_m",
    "tfn_u",
    "submission_timestamp",
]


def completed_responses_dataframe(
    responses: Sequence[DistributedResponseRecord],
    assignments: Sequence[AssignmentRecord],
) -> pd.DataFrame:
    """Return analysis-ready responses from completed respondents only."""

    completed_ids = {
        assignment["respondent_id"]
        for assignment in assignments
        if assignment["status"] == "completed"
    }
    rows = [
        {
            "respondent_id": response["submission_id"],
            "expert_code": response["expert_code"],
            "set_id": int(response["set_id"]),
            "source_variable_code": response["from_factor"],
            "source_variable_name": response["source_variable_name"],
            "target_variable_code": response["to_factor"],
            "target_variable_name": response["target_variable_name"],
            "linguistic_response": response["linguistic_value"],
            "tfn_l": response["tfn_l"],
            "tfn_m": response["tfn_m"],
            "tfn_u": response["tfn_u"],
            "submission_timestamp": response["timestamp"],
        }
        for response in responses
        if response["submission_id"] in completed_ids
    ]
    frame = pd.DataFrame.from_records(rows, columns=ADMIN_RESPONSE_COLUMNS)
    if frame.empty:
        return frame
    factor_order = {code: index for index, code in enumerate(FACTOR_CODES)}
    frame["_source_order"] = frame["source_variable_code"].map(factor_order)
    frame["_target_order"] = frame["target_variable_code"].map(factor_order)
    return (
        frame.sort_values(
            [
                "set_id",
                "respondent_id",
                "_source_order",
                "_target_order",
            ]
        )
        .drop(columns=["_source_order", "_target_order"])
        .reset_index(drop=True)
    )


def relationship_coverage_dataframe(
    completed_responses: pd.DataFrame,
    *,
    minimum_evaluations: int,
) -> pd.DataFrame:
    """Return all 306 relationships with total and usable evaluation counts."""

    if completed_responses.empty:
        total_counts: dict[tuple[str, str], int] = {}
        cannot_counts: dict[tuple[str, str], int] = {}
    else:
        pair_columns = ["source_variable_code", "target_variable_code"]
        total_counts = completed_responses.groupby(pair_columns).size().to_dict()
        cannot_counts = (
            completed_responses[
                completed_responses["linguistic_response"]
                == CANNOT_ASSESS_VALUE
            ]
            .groupby(pair_columns)
            .size()
            .to_dict()
        )

    rows = []
    for relationship in all_relationships():
        pair = (relationship.source_code, relationship.target_code)
        total = int(total_counts.get(pair, 0))
        cannot_assess = int(cannot_counts.get(pair, 0))
        usable = total - cannot_assess
        rows.append(
            {
                "set_id": relationship.set_id,
                "source_variable_code": relationship.source_code,
                "source_variable_name": relationship.source_name,
                "target_variable_code": relationship.target_code,
                "target_variable_name": relationship.target_name,
                "evaluation_count": total,
                "cannot_assess_count": cannot_assess,
                "usable_evaluation_count": usable,
                "minimum_required": minimum_evaluations,
                "enough_evaluations": usable >= minimum_evaluations,
            }
        )
    return pd.DataFrame.from_records(rows)


def set_summary_dataframe(
    assignments: Sequence[AssignmentRecord],
) -> pd.DataFrame:
    """Return assignment and completion counts for all seven sets."""

    rows = []
    for set_id in range(1, 8):
        set_assignments = [
            assignment
            for assignment in assignments
            if int(assignment["set_id"]) == set_id
        ]
        completed = sum(
            assignment["status"] == "completed"
            for assignment in set_assignments
        )
        rows.append(
            {
                "set_id": set_id,
                "assigned_respondents": len(set_assignments),
                "completed_respondents": completed,
                "in_progress_respondents": len(set_assignments) - completed,
            }
        )
    return pd.DataFrame.from_records(rows)


def _count_matrix(coverage: pd.DataFrame) -> pd.DataFrame:
    matrix = coverage.pivot(
        index="source_variable_code",
        columns="target_variable_code",
        values="usable_evaluation_count",
    ).reindex(index=FACTOR_CODES, columns=FACTOR_CODES)
    for code in FACTOR_CODES:
        matrix.loc[code, code] = 0
    matrix.index.name = "source_variable_code"
    matrix.columns.name = "target_variable_code"
    return matrix.astype(int)


def generate_administrator_exports(
    responses: Sequence[DistributedResponseRecord],
    assignments: Sequence[AssignmentRecord],
    *,
    minimum_evaluations: int,
) -> AdministratorExportBundle:
    """Generate combined CSV and Excel files for matrix reconstruction."""

    completed = completed_responses_dataframe(responses, assignments)
    coverage = relationship_coverage_dataframe(
        completed,
        minimum_evaluations=minimum_evaluations,
    )
    set_summary = set_summary_dataframe(assignments)
    metadata = pd.DataFrame(
        {
            "field": [
                "matrix_orientation",
                "directed_relationships",
                "diagonal_rule",
                "questionnaire_sets",
                "cannot_assess_rule",
                "aggregation_status",
            ],
            "value": [
                "Source variable influences target variable",
                "306",
                "Diagonal is fixed at zero and is never presented",
                "7 disjoint balanced sets",
                "TFN fields are blank and excluded from usable counts",
                "Raw evaluations only; no Fuzzy DEMATEL aggregation performed",
            ],
        }
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        completed.to_excel(writer, sheet_name="Responses_Long", index=False)
        coverage.to_excel(writer, sheet_name="Relationship_Coverage", index=False)
        _count_matrix(coverage).to_excel(
            writer, sheet_name="Evaluation_Count_Matrix", index=True
        )
        set_summary.to_excel(writer, sheet_name="Set_Summary", index=False)
        _definitions_dataframe().to_excel(
            writer, sheet_name="Factor_Definitions", index=False
        )
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        _format_workbook(writer)

    return AdministratorExportBundle(
        responses_csv=_dataframe_to_csv_bytes(completed, index=False),
        complete_excel=buffer.getvalue(),
    )


HIERARCHICAL_RESPONSE_COLUMNS = [
    "respondent_id",
    "expert_code",
    "matrix_id",
    "matrix_name",
    "source_code",
    "source_name",
    "target_code",
    "target_name",
    "linguistic_response",
    "tfn_l",
    "tfn_m",
    "tfn_u",
    "submission_timestamp",
]


def completed_hierarchical_responses_dataframe(
    responses: Sequence[HierarchicalResponseRecord],
    questionnaires: Sequence[HierarchicalQuestionnaireRecord],
) -> pd.DataFrame:
    """Return all 104 raw answers for each completed hierarchical respondent."""

    completed_ids = {
        questionnaire["respondent_id"]
        for questionnaire in questionnaires
        if questionnaire["status"] == "completed"
    }
    matrix_names = {
        matrix.id: matrix.label for matrix in matrix_definitions()
    }
    rows = [
        {
            "respondent_id": response["respondent_id"],
            "expert_code": response["expert_code"],
            "matrix_id": response["matrix_id"],
            "matrix_name": matrix_names[response["matrix_id"]],
            "source_code": response["source_code"],
            "source_name": response["source_name"],
            "target_code": response["target_code"],
            "target_name": response["target_name"],
            "linguistic_response": response["linguistic_value"],
            "tfn_l": response["tfn_l"],
            "tfn_m": response["tfn_m"],
            "tfn_u": response["tfn_u"],
            "submission_timestamp": response["responded_at"],
        }
        for response in responses
        if response["respondent_id"] in completed_ids
    ]
    frame = pd.DataFrame.from_records(rows, columns=HIERARCHICAL_RESPONSE_COLUMNS)
    if frame.empty:
        return frame
    relationship_order = {
        relationship.key: index
        for index, relationship in enumerate(all_hierarchical_relationships())
    }
    frame["_relationship_order"] = frame.apply(
        lambda row: relationship_order[
            f"{row['matrix_id']}|{row['source_code']}|{row['target_code']}"
        ],
        axis=1,
    )
    return (
        frame.sort_values(["respondent_id", "_relationship_order"])
        .drop(columns="_relationship_order")
        .reset_index(drop=True)
    )


def hierarchical_coverage_dataframe(
    completed_responses: pd.DataFrame,
    *,
    minimum_evaluations: int,
) -> pd.DataFrame:
    """Return all 104 allowed relationships with completed evaluation counts."""

    if completed_responses.empty:
        counts: dict[tuple[str, str, str], int] = {}
    else:
        counts = (
            completed_responses.groupby(
                ["matrix_id", "source_code", "target_code"]
            )
            .size()
            .to_dict()
        )
    return pd.DataFrame.from_records(
        {
            "matrix_id": relationship.matrix_id,
            "matrix_name": relationship.matrix_label,
            "source_code": relationship.source_code,
            "source_name": relationship.source_name,
            "target_code": relationship.target_code,
            "target_name": relationship.target_name,
            "evaluation_count": int(
                counts.get(
                    (
                        relationship.matrix_id,
                        relationship.source_code,
                        relationship.target_code,
                    ),
                    0,
                )
            ),
            "minimum_required": minimum_evaluations,
            "enough_evaluations": int(
                counts.get(
                    (
                        relationship.matrix_id,
                        relationship.source_code,
                        relationship.target_code,
                    ),
                    0,
                )
            )
            >= minimum_evaluations,
        }
        for relationship in all_hierarchical_relationships()
    )


def hierarchical_respondent_summary_dataframe(
    questionnaires: Sequence[HierarchicalQuestionnaireRecord],
) -> pd.DataFrame:
    """Return one anonymous status row per hierarchical respondent."""

    return pd.DataFrame.from_records(
        (
            {
                "respondent_id": item["respondent_id"],
                "expert_code": item["expert_code"],
                "design_version": item["design_version"],
                "status": item["status"],
                "started_at": item["started_at"],
                "completed_at": item["completed_at"],
            }
            for item in questionnaires
        ),
        columns=[
            "respondent_id",
            "expert_code",
            "design_version",
            "status",
            "started_at",
            "completed_at",
        ],
    )


def hierarchical_matrix_summary_dataframe(
    completed_responses: pd.DataFrame,
) -> pd.DataFrame:
    """Summarize the fixed pair count and collected answers by matrix."""

    return pd.DataFrame.from_records(
        {
            "matrix_id": matrix.id,
            "matrix_name": matrix.label,
            "matrix_size": f"{len(matrix.criteria)} × {len(matrix.criteria)}",
            "required_per_respondent": matrix.required_comparisons,
            "completed_evaluations_collected": (
                0
                if completed_responses.empty
                else int(
                    (completed_responses["matrix_id"] == matrix.id).sum()
                )
            ),
        }
        for matrix in matrix_definitions()
    )


def hierarchical_wide_responses_dataframe(
    completed_responses: pd.DataFrame,
) -> pd.DataFrame:
    """Return one respondent per row with all 104 linguistic directions."""

    relationship_columns = [
        relationship.key for relationship in all_hierarchical_relationships()
    ]
    identity_columns = ["respondent_id", "expert_code"]
    if completed_responses.empty:
        return pd.DataFrame(columns=identity_columns + relationship_columns)
    working = completed_responses.copy()
    working["relationship_key"] = (
        working["matrix_id"]
        + "|"
        + working["source_code"]
        + "|"
        + working["target_code"]
    )
    wide = working.pivot(
        index=identity_columns,
        columns="relationship_key",
        values="linguistic_response",
    ).reindex(columns=relationship_columns)
    wide.columns.name = None
    return wide.reset_index()


def _hierarchical_count_matrix(
    coverage: pd.DataFrame, matrix_id: str
) -> pd.DataFrame:
    matrix = get_matrix_definition(matrix_id)
    codes = [criterion.code for criterion in matrix.criteria]
    subset = coverage[coverage["matrix_id"] == matrix_id]
    wide = subset.pivot(
        index="source_code", columns="target_code", values="evaluation_count"
    ).reindex(index=codes, columns=codes)
    for code in codes:
        wide.loc[code, code] = 0
    wide.index.name = "source_code"
    wide.columns.name = "target_code"
    return wide.fillna(0).astype(int)


def _hierarchical_definitions_dataframe() -> pd.DataFrame:
    return pd.DataFrame.from_records(
        {
            "matrix_id": matrix.id,
            "matrix_name": matrix.label,
            "code": criterion.code,
            "name": criterion.name,
            "definition": criterion.definition,
        }
        for matrix in matrix_definitions()
        for criterion in matrix.criteria
    )


def generate_hierarchical_administrator_exports(
    responses: Sequence[HierarchicalResponseRecord],
    questionnaires: Sequence[HierarchicalQuestionnaireRecord],
    *,
    minimum_evaluations: int,
) -> AdministratorExportBundle:
    """Generate analysis-ready CSV and Excel for the four-matrix design."""

    completed = completed_hierarchical_responses_dataframe(
        responses, questionnaires
    )
    coverage = hierarchical_coverage_dataframe(
        completed, minimum_evaluations=minimum_evaluations
    )
    respondent_summary = hierarchical_respondent_summary_dataframe(questionnaires)
    matrix_summary = hierarchical_matrix_summary_dataframe(completed)
    wide_responses = hierarchical_wide_responses_dataframe(completed)
    metadata = pd.DataFrame(
        {
            "field": [
                "questionnaire_design",
                "matrix_orientation",
                "required_evaluations_per_respondent",
                "matrix_pair_counts",
                "diagonal_rule",
                "cross_dimension_criterion_rule",
                "aggregation_status",
            ],
            "value": [
                "Hierarchical Fuzzy DEMATEL (hierarchical_v1)",
                "ROW/source influences COLUMN/target",
                "104",
                "cultural=30; economic=12; strategic=56; dimension_level=6",
                "Diagonal is fixed at zero and is never a respondent answer",
                "Individual criteria from different dimensions are not evaluated",
                "Raw evaluations only; no Fuzzy DEMATEL calculations performed",
            ],
        }
    )
    count_sheet_names = {
        "cultural": "Cultural_Counts",
        "economic": "Economic_Counts",
        "strategic": "Strategic_Counts",
        "dimension_level": "Dimension_Counts",
    }

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        completed.to_excel(writer, sheet_name="Responses_Long", index=False)
        wide_responses.to_excel(writer, sheet_name="Responses_Wide", index=False)
        coverage.to_excel(writer, sheet_name="Relationship_Coverage", index=False)
        respondent_summary.to_excel(
            writer, sheet_name="Respondent_Summary", index=False
        )
        matrix_summary.to_excel(writer, sheet_name="Matrix_Summary", index=False)
        for matrix_id, sheet_name in count_sheet_names.items():
            _hierarchical_count_matrix(coverage, matrix_id).to_excel(
                writer, sheet_name=sheet_name, index=True
            )
        _hierarchical_definitions_dataframe().to_excel(
            writer, sheet_name="Criteria_Definitions", index=False
        )
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        _format_workbook(writer)

    return AdministratorExportBundle(
        responses_csv=_dataframe_to_csv_bytes(completed, index=False),
        complete_excel=buffer.getvalue(),
    )
